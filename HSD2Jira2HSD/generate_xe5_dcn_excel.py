#!/usr/bin/env python3
"""
Generate xe5 DCN Enabling Status Excel.

Writes:  xe5 DCN enabling status.xlsx   (same folder as this script)

Two-sheet structure matching the JGS "GT DCNs" approach:
  Sheet 1 - "xe5 GT DCNs"      formula-based view (Dev Readiness + Impl. Trend)
  Sheet 2 - "DO NOT MODIFY"    raw data fetched from HSD + Jira

Scoping flags (new):
  "Pending arch scoping"        arch sw_impact child has to_be_assigned exposure
  "Pending engineering scoping" arch is scoped but dev child still has to_be_assigned
HSD Status + HSD State columns added to the main sheet.

Usage:
    cd /path/to/HSD2Jira2HSD
    python3 generate_xe5_dcn_excel.py
"""
import logging
import sys
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import requests
import urllib3
from requests_kerberos import HTTPKerberosAuth, OPTIONAL
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-7s  %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger('xe5dcn')

# ── Constants ─────────────────────────────────────────────────────────────────
BASE_DIR        = Path(__file__).resolve().parent
TOKEN_PATH      = BASE_DIR / '.jira_token'
OUTPUT_PATH     = BASE_DIR / 'xe5 DCN enabling status.xlsx'

XE5_QUERY_ID    = '14028312865'
HSD_BASE        = 'https://hsdes.intel.com'
HSD_REST        = HSD_BASE + '/rest/article'
ESSERVICE_URL   = HSD_BASE + '/ws/ESService'
ESSERVICE_HDRS  = {'APP': 'HSD-ES Article', 'Accept': 'application/json',
                   'Content-Type': 'application/json'}
JIRA_API        = 'https://jira.devtools.intel.com/rest/api/2'
JIRA_TREND_WW   = 'customfield_34504'
HSD_ARTICLE_URL = 'https://hsdes.intel.com/appstore/article-one/#/article/'
JIRA_BROWSE_URL = 'https://jira.devtools.intel.com/browse/'

# (dev sw_component keyword, column key, display label)
DEV_COMPONENTS = [
    ('compute',        'compute', 'Compute'),
    ('igc_compute',    'igc',     'IGC Compute'),
    ('i915_kmd',       'kmd',     'XeKMD'),
    ('pisa_finaliser', 'pisa',    'PISA Finaliser'),
    ('uc_global',      'guc',     'GUC'),
]

# Architecture sw_component keyword that gates each dev component.
DEV_ARCH_KW = {
    'compute': 'compute',   # compute/arch  -> compute/dev
    'igc':     'compiler',  # compiler/arch -> igc_compute/dev
    'kmd':     'core',      # core/arch     -> i915_kmd/dev
    'pisa':    'compiler',  # compiler/arch -> pisa_finaliser/dev
    'guc':     'guc',       # guc/arch      -> uc_global/dev
}

NO_SCOPE = ('', 'none', 'to_be_assigned')  # exposure values meaning "not yet scoped"

# DATA sheet column map: key -> (jira_col, status_col, trend_col, done_col)
DATA_COL = {
    'compute': ('E', 'F', 'G', 'H'),
    'igc':     ('I', 'J', 'K', 'L'),
    'kmd':     ('M', 'N', 'O', 'P'),
    'pisa':    ('Q', 'R', 'S', 'T'),
    'guc':     ('U', 'V', 'W', 'X'),
}
DATA_COL_IDX = {
    'compute': (5, 6, 7, 8),
    'igc':     (9, 10, 11, 12),
    'kmd':     (13, 14, 15, 16),
    'pisa':    (17, 18, 19, 20),
    'guc':     (21, 22, 23, 24),
}
BLANK = {'jira': '(blank)', 'status': '(blank)', 'trend': '(blank)', 'done': '(blank)'}


# ── HSD helpers ───────────────────────────────────────────────────────────────

def new_session():
    s = requests.Session()
    s.auth   = HTTPKerberosAuth(mutual_authentication=OPTIONAL)
    s.verify = False
    return s


def esservice(session, command, command_args):
    payload = {'requests': [{'api_client': 'HSD-ES Article',
                              'tran_id': str(uuid.uuid4()).upper(),
                              'command': command,
                              'command_args': command_args,
                              'var_args': [], 'copy_args': []}]}
    try:
        r = session.post(ESSERVICE_URL, headers=ESSERVICE_HDRS, json=payload, timeout=30)
        if r.status_code == 200:
            resp0 = r.json().get('responses', [{}])[0]
            if resp0.get('status') == 'success':
                return resp0.get('result_table', [])
            log.warning('ESService non-success: %s', resp0.get('message'))
    except Exception as exc:
        log.error('ESService %s failed: %s', command, exc)
    return None


def execute_query(session, query_id):
    result = esservice(session, 'execute_saved_query',
                       {'query_id': str(query_id), 'fields': ['id']})
    if result:
        ids = [str(r['id']) for r in result if isinstance(r, dict) and r.get('id')]
        if ids:
            return ids
    for url, params in [
        (f'{HSD_BASE}/rest/query/{query_id}', {'fields': 'id', 'count': 500}),
        (f'{HSD_BASE}/rest/article',          {'query_id': query_id, 'fields': 'id', 'count': 500}),
    ]:
        try:
            r = session.get(url, params=params, timeout=30)
            if r.status_code == 200:
                data  = r.json()
                items = data.get('data', []) if isinstance(data, dict) else (data or [])
                ids   = [str(i['id']) for i in items if isinstance(i, dict) and i.get('id')]
                if ids:
                    return ids
        except Exception as exc:
            log.warning('query fallback: %s', exc)
    return []


def fetch_hsd_info(session, hsd_id):
    """Return (title, hsd_status, hsd_state) for an HSD article."""
    try:
        r = session.get(f'{HSD_REST}/{hsd_id}', timeout=20)
        if r.status_code == 200:
            items = r.json().get('data', [])
            if items:
                title  = str(items[0].get('title')  or '').strip() or f'HSD {hsd_id}'
                status = str(items[0].get('status') or '').strip()
                # Try several field names HSD uses for 'state'
                state = (
                    str(items[0].get('state')                          or '').strip() or
                    str(items[0].get('ip_hw_graphics.bugeco.state')    or '').strip() or
                    str(items[0].get('ip_hw_graphics.feature.state')   or '').strip() or
                    str(items[0].get('classification')                 or '').strip()
                )
                return title, status, state
    except Exception as exc:
        log.error('HSD REST %s: %s', hsd_id, exc)
    return f'HSD {hsd_id}', '', ''


def get_sw_impact_children(session, hsd_id):
    links = esservice(session, 'get_related_records', {'id': str(hsd_id)})
    if not links:
        return []
    sw_ids = [
        str(r['id'])
        for r in links
        if isinstance(r, dict)
        and str(r.get('subject', '')).lower() == 'sw_impact'
        and 'ip_hw_graphics' in str(r.get('tenant', '')).lower().replace('.', '_').replace('-', '_')
        and r.get('id')
    ]
    records = []
    for sw_id in sw_ids:
        rec = esservice(session, 'get_record_by_id', {'id': sw_id})
        if rec and isinstance(rec, list) and rec[0]:
            records.append(rec[0])
        elif isinstance(rec, dict):
            records.append(rec)
    return records


def parse_sw_impact(rec):
    def g(*keys):
        for k in keys:
            v = str(rec.get(k) or '').strip()
            if v and v.lower() not in ('none', 'null', ''):
                return v
        return ''
    return {
        'sw_component': g('ip_hw_graphics.sw_impact.sw_component', 'sw_component'),
        'sw_task':      g('ip_hw_graphics.sw_impact.sw_task',      'sw_task'),
        'sw_exposure':  g('ip_hw_graphics.sw_impact.sw_exposure',  'sw_exposure'),
        'sw_record':    g('ip_hw_graphics.sw_impact.sw_record',    'sw_record'),
        'done':         g('ip_hw_graphics.sw_impact.done',         'done'),
    }


def fetch_jira(token, key):
    hdrs = {'Authorization': f'Bearer {token}', 'Accept': 'application/json'}
    try:
        r = requests.get(f'{JIRA_API}/issue/{key}',
                         params={'fields': f'status,{JIRA_TREND_WW}'},
                         headers=hdrs, verify=False, timeout=20)
        if not r.ok:
            return '', ''
        fields = r.json().get('fields', {})
        status = (fields.get('status') or {}).get('name', '') or ''
        trend  = str(fields.get(JIRA_TREND_WW) or '').strip().split('.')[0]
        return status.strip(), trend
    except Exception as exc:
        log.error('Jira fetch %s: %s', key, exc)
        return '', ''


# ── Per-parent fetch ──────────────────────────────────────────────────────────

def process_parent(token, hsd_id):
    """Fetch one parent DCN. Creates its own HSD session (thread-safe)."""
    session = new_session()
    try:
        title, hsd_status, hsd_state = fetch_hsd_info(session, hsd_id)
        log.info('  -> %s  %s', hsd_id, title[:70])

        sw_raw    = get_sw_impact_children(session, hsd_id)
        sw_parsed = [parse_sw_impact(r) for r in sw_raw]

        comp_data = {}
        for sw_kw, key, _ in DEV_COMPONENTS:
            arch_kw = DEV_ARCH_KW[key]

            # Dev children with valid exposure OR already marked done
            dev_valid = [
                c for c in sw_parsed
                if sw_kw in c['sw_component'].lower()
                and c['sw_task'].lower() == 'development'
                and (c['sw_exposure'].lower() not in NO_SCOPE
                     or c.get('done', '').lower() in ('yes', 'true', '1'))
            ]

            if dev_valid:
                sw_record = dev_valid[0]['sw_record'].strip()
                if not sw_record:
                    comp_data[key] = {**BLANK,
                                      'status': f"No Jira ({dev_valid[0]['sw_exposure']})"}
                else:
                    jira_status, trend_ww = fetch_jira(token, sw_record)
                    # Mark done if Jira is closed OR sw_impact done field is set
                    sw_done   = dev_valid[0].get('done', '').lower() in ('yes', 'true', '1')
                    jira_done = jira_status.lower() in ('closed', 'done', 'resolved', 'implemented')
                    done = 'yes' if (sw_done or jira_done) else 'no'
                    comp_data[key] = {'jira':   sw_record,
                                      'status': jira_status or '(blank)',
                                      'trend':  trend_ww    or '(blank)',
                                      'done':   done}
                continue

            # No valid dev child -- check architecture gate
            arch_children = [
                c for c in sw_parsed
                if arch_kw in c['sw_component'].lower()
                and c['sw_task'].lower() == 'architecture'
            ]
            if not arch_children:
                comp_data[key] = dict(BLANK)
                continue

            arch_exp = arch_children[0]['sw_exposure'].lower()
            if arch_exp == 'to_be_assigned':
                # Arch gate explicitly not yet assessed
                comp_data[key] = {**BLANK, 'status': 'Pending arch scoping'}
            elif arch_exp in ('', 'none'):
                # Arch explicitly said not involved -- show ---
                comp_data[key] = dict(BLANK)
            else:
                # Arch scoped -- flag ONLY if dev child EXISTS but is stuck at to_be_assigned.
                # No dev child at all = architecture done, engineering not engaged yet = show ---.
                dev_pending = [
                    c for c in sw_parsed
                    if sw_kw in c['sw_component'].lower()
                    and c['sw_task'].lower() == 'development'
                    and c['sw_exposure'].lower() == 'to_be_assigned'
                ]
                if dev_pending:
                    comp_data[key] = {**BLANK, 'status': 'Pending engineering scoping'}
                else:
                    comp_data[key] = dict(BLANK)

        has_kmd = any(
            'i915_kmd' in c['sw_component'].lower()
            and c['sw_task'].lower() == 'development'
            and c['sw_exposure'].lower() not in NO_SCOPE
            for c in sw_parsed)
        has_compute = any(
            c['sw_component'].lower() == 'compute'
            and c['sw_task'].lower() == 'development'
            and c['sw_exposure'].lower() not in NO_SCOPE
            for c in sw_parsed)

        return {'id': hsd_id, 'title': title,
                'hsd_status': hsd_status, 'hsd_state': hsd_state,
                'comp_data': comp_data,
                'kmd_dcn': 'yes' if has_kmd else 'no',
                'compute_dcn': 'yes' if has_compute else 'no'}

    except Exception as exc:
        log.error('Error on HSD %s: %s', hsd_id, exc)
        return {'id': hsd_id, 'title': f'HSD {hsd_id}',
                'hsd_status': '', 'hsd_state': '',
                'comp_data': {}, 'kmd_dcn': 'no', 'compute_dcn': 'no'}


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(rows):
    wb   = Workbook()
    _ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _wrap = Alignment(vertical='center', wrap_text=True)

    # Sheet: DO NOT MODIFY
    # A-D: metadata  E-X: 5 components x 4 cols  Y:KMD DCN  Z:Compute DCN  AA:hsd_state
    data_ws = wb.active
    data_ws.title = 'DO NOT MODIFY'

    DATA_HEADERS = [
        'ord_num', 'id', 'Title', 'hsd_status',
        'dev_compute_jira', 'dev_compute_status', 'dev_compute_trend', 'dev_compute_done',
        'dev_igc_jira',     'dev_igc_status',     'dev_igc_trend',     'dev_igc_done',
        'dev_kmd_jira',     'dev_kmd_status',     'dev_kmd_trend',     'dev_kmd_done',
        'dev_pisa_jira',    'dev_pisa_status',    'dev_pisa_trend',    'dev_pisa_done',
        'dev_guc_jira',     'dev_guc_status',     'dev_guc_trend',     'dev_guc_done',
        'KMD DCN', 'Compute DCN', 'hsd_state',
    ]
    hdr_fill = PatternFill('solid', fgColor='2C4770')
    hdr_font = Font(color='FFFFFF', bold=True, size=9)
    for c, h in enumerate(DATA_HEADERS, 1):
        cell = data_ws.cell(1, c, h)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = _ctr

    for i, row in enumerate(rows, 1):
        dr = i + 1
        cd = row['comp_data']
        data_ws.cell(dr, 1, i)
        data_ws.cell(dr, 2, row['id'])
        data_ws.cell(dr, 3, row['title'])
        data_ws.cell(dr, 4, row.get('hsd_status', ''))
        for key, (jc, sc, tc, dc) in DATA_COL_IDX.items():
            d = cd.get(key, BLANK)
            data_ws.cell(dr, jc, d['jira'])
            data_ws.cell(dr, sc, d['status'])
            data_ws.cell(dr, tc, d['trend'])
            data_ws.cell(dr, dc, d['done'])
        data_ws.cell(dr, 25, row['kmd_dcn'])
        data_ws.cell(dr, 26, row['compute_dcn'])
        data_ws.cell(dr, 27, row.get('hsd_state', ''))  # AA
    data_ws.column_dimensions['A'].width = 8
    data_ws.column_dimensions['B'].width = 14
    data_ws.column_dimensions['C'].width = 52
    data_ws.column_dimensions['D'].width = 14
    for c in range(5, 28):
        data_ws.column_dimensions[get_column_letter(c)].width = 15
    data_ws.freeze_panes = 'C2'

    # Sheet: xe5 GT DCNs (19 cols)
    # A:ID B:Title C:Enabling D:Compute-inv E:KMD-inv F:HSD-Status G:HSD-State
    # H-L: Dev Readiness   M-Q: Trends   R: Impl.Trend   S: Comments
    main_ws = wb.create_sheet('xe5 GT DCNs', 0)
    DN = "'DO NOT MODIFY'"

    main_ws.cell(1, 1,
        f'xe5 DCN Enabling Status -- {len(rows)} DCNs -- '
        f'Generated {datetime.utcnow().strftime("%Y-%m-%d")}')
    main_ws.merge_cells('A1:S1')
    main_ws['A1'].fill      = PatternFill('solid', fgColor='1F3864')
    main_ws['A1'].font      = Font(color='FFFFFF', bold=True, size=12)
    main_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    main_ws.row_dimensions[1].height = 26

    for rng, label, fg in [
        ('A2:G2', 'xe5 DCN Info',          'C9D7F0'),
        ('H2:L2', 'Development Readiness', 'D6E4BC'),
        ('M2:R2', 'Implementation Trends', 'FCE4D6'),
        ('S2:S2', 'Comments',              'F2F2F2'),
    ]:
        main_ws.merge_cells(rng)
        c = main_ws[rng.split(':')[0]]
        c.value = label
        c.fill  = PatternFill('solid', fgColor=fg)
        c.font  = Font(bold=True, size=10)
        c.alignment = _ctr
    main_ws.row_dimensions[2].height = 20

    col_hdr_fill = PatternFill('solid', fgColor='4472C4')
    for c, h in enumerate([
        'ID', 'Title', 'Enabling', 'Compute involved', 'KMD involved',
        'HSD Status', 'HSD State',
        'Compute', 'IGC Compute', 'XeKMD', 'PISA', 'GUC',
        'Compute', 'IGC Compute', 'XeKMD', 'PISA', 'GUC', 'Impl. Trend',
        'Comments',
    ], 1):
        cell = main_ws.cell(3, c, h)
        cell.fill = col_hdr_fill
        cell.font = Font(color='FFFFFF', bold=True, size=9)
        cell.alignment = _ctr
    main_ws.row_dimensions[3].height = 30

    def dev_f(jc, sc, dc, mr, dr):
        return (
            f"=IF({DN}!{jc}{dr}<>\"\"," +
            f"IF({DN}!{jc}{dr}=\"(blank)\"," +
            f"IF({DN}!{sc}{dr}=\"(blank)\",\"--\",{DN}!{sc}{dr})," +
            f"IF({DN}!{dc}{dr}=\"yes\"," +
            f"HYPERLINK(\"{JIRA_BROWSE_URL}\"&{DN}!{jc}{dr},\"Done\")," +
            f"HYPERLINK(\"{JIRA_BROWSE_URL}\"&{DN}!{jc}{dr},{DN}!{sc}{dr})))," +
            f"IF(A{mr}<>\"\",\"In Analysis\",\"\"))"
        )

    def trend_f(tc, mr, dr):
        return (
            f"=IF({DN}!{tc}{dr}<>\"\"," +
            f"IF({DN}!{tc}{dr}=\"(blank)\",\"--\",{DN}!{tc}{dr})," +
            f"IF(A{mr}<>\"\",\"In Analysis\",\"\"))"
        )

    def impl_trend_f(mr):
        def ww(col):
            v = f'{col}{mr}'
            return (f'IF(ISNUMBER(VALUE(RIGHT({v},2))),' +
                    f'VALUE(RIGHT(_xlfn.TEXTBEFORE({v},"ww"),2)&_xlfn.TEXTAFTER({v},"ww")),0)')
        mx = 'MAX(' + ','.join(ww(c) for c in ['M', 'N', 'O', 'P', 'Q']) + ')'
        return (f'=IF({mx}=0,' +
                f'IF(A{mr}<>"",IF(B{mr}<>"<DCN removed>","In Analysis","--"),""),' +
                f'LEFT({mx},2)&"ww"&RIGHT({mx},2))')

    id_font = Font(color='0563C1', underline='single')
    for i in range(len(rows)):
        mr = i + 4
        dr = i + 2
        main_ws.cell(mr, 1,
            f"=IF({DN}!B{dr}<>\"\"," +
            f"HYPERLINK(\"{HSD_ARTICLE_URL}\"&{DN}!B{dr},{DN}!B{dr}),\"\")"
        ).font = id_font
        main_ws.cell(mr, 1).alignment = _ctr
        main_ws.cell(mr, 2, f"=IF({DN}!C{dr}<>\"\",{DN}!C{dr},\"\")").alignment = _wrap
        main_ws.cell(mr, 4, f"=IF({DN}!Z{dr}<>\"\",{DN}!Z{dr},\"\")").alignment = _ctr
        main_ws.cell(mr, 5, f"=IF({DN}!Y{dr}<>\"\",{DN}!Y{dr},\"\")").alignment = _ctr
        main_ws.cell(mr, 6, f"=IF({DN}!D{dr}<>\"\",{DN}!D{dr},\"\")").alignment = _ctr
        main_ws.cell(mr, 7, f"=IF({DN}!AA{dr}<>\"\",{DN}!AA{dr},\"\")").alignment = _ctr
        for off, (jc, sc, dc) in enumerate([
            (DATA_COL['compute'][0], DATA_COL['compute'][1], DATA_COL['compute'][3]),
            (DATA_COL['igc'][0],     DATA_COL['igc'][1],     DATA_COL['igc'][3]),
            (DATA_COL['kmd'][0],     DATA_COL['kmd'][1],     DATA_COL['kmd'][3]),
            (DATA_COL['pisa'][0],    DATA_COL['pisa'][1],    DATA_COL['pisa'][3]),
            (DATA_COL['guc'][0],     DATA_COL['guc'][1],     DATA_COL['guc'][3]),
        ]):
            main_ws.cell(mr, 8 + off, dev_f(jc, sc, dc, mr, dr)).alignment = _ctr
        for off, tc in enumerate([
            DATA_COL['compute'][2], DATA_COL['igc'][2], DATA_COL['kmd'][2],
            DATA_COL['pisa'][2],    DATA_COL['guc'][2],
        ]):
            main_ws.cell(mr, 13 + off, trend_f(tc, mr, dr)).alignment = _ctr
        main_ws.cell(mr, 18, impl_trend_f(mr)).alignment = _ctr
        main_ws.row_dimensions[mr].height = 18

    for c, w in enumerate([12, 58, 10, 14, 12, 14, 12, 12, 14, 10, 14, 10, 12, 14, 10, 14, 10, 12, 35], 1):
        main_ws.column_dimensions[get_column_letter(c)].width = w
    main_ws.freeze_panes = 'C4'
    main_ws.auto_filter.ref = f'A3:{get_column_letter(19)}3'

    # ── Conditional formatting ────────────────────────────────────────────────
    from openpyxl.formatting.rule import FormulaRule
    n = len(rows)
    dev_rng   = f'H4:L{3 + n}'
    trend_rng = f'M4:Q{3 + n}'

    # Done = green cell background + dark-green bold text
    # NOTE: DXF conditional-format fills use bgColor (not fgColor) for the rendered cell background
    main_ws.conditional_formatting.add(dev_rng, FormulaRule(
        formula=['H4="Done"'],
        fill=PatternFill(bgColor='C6EFCE'),
        font=Font(color='276221', bold=True)))

    # Pending arch scoping = amber
    main_ws.conditional_formatting.add(dev_rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("arch scoping",H4))'],
        fill=PatternFill(bgColor='FFEB9C'),
        font=Font(color='9C6500', bold=True)))

    # Pending engineering scoping = light orange
    main_ws.conditional_formatting.add(dev_rng, FormulaRule(
        formula=['ISNUMBER(SEARCH("engineering scoping",H4))'],
        fill=PatternFill(bgColor='FCE4D6'),
        font=Font(color='843C0C', bold=True)))

    return wb


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not TOKEN_PATH.exists():
        log.error('Jira token not found at %s', TOKEN_PATH)
        sys.exit(1)
    token = TOKEN_PATH.read_text().strip()
    if not token:
        log.error('Jira token file is empty')
        sys.exit(1)
    log.info('Reading HSD query %s ...', XE5_QUERY_ID)
    session    = new_session()
    parent_ids = execute_query(session, XE5_QUERY_ID)
    if not parent_ids:
        log.error('No parent DCNs returned -- check Kerberos ticket (kinit)')
        sys.exit(1)
    log.info('%d parent DCNs found', len(parent_ids))
    log.info('Fetching sw_impact children + Jira status (parallel, 6 workers) ...')
    rows = []
    with ThreadPoolExecutor(max_workers=6) as pool:
        futs = {pool.submit(process_parent, token, hid): hid for hid in parent_ids}
        done = 0
        for fut in as_completed(futs):
            rows.append(fut.result())
            done += 1
            if done % 10 == 0:
                log.info('  %d / %d done', done, len(parent_ids))
    rows.sort(key=lambda x: x['id'])
    log.info('Building Excel ...')
    wb = build_excel(rows)
    wb.save(OUTPUT_PATH)
    log.info('Saved -> %s', OUTPUT_PATH)


if __name__ == '__main__':
    main()

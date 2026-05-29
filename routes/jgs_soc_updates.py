"""
JGS SOC Updates — /jgs-soc-updates

Reads task data from JGS_SOC_Trend.xlsx (via soc_automation directory),
builds graph/filter data, and renders the dashboard.
"""
import io
import json
import re as _re
import sqlite3
import ssl
import logging
import urllib.request
import urllib.parse
import urllib.error
from collections import Counter
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import requests as _requests
from flask import Blueprint, render_template, request, jsonify

from config import SOC_AUTOMATION_DIR, CACHE_TTL, JIRA_TOKEN_PATH
from utils.helpers import parse_ww, is_cell_green, extract_jira_keys, current_ww_info

logger = logging.getLogger(__name__)
bp = Blueprint('jgs_soc', __name__)


def _graph_category(cat: str) -> str:
    """Normalise category name for the graph: all UAL* sub-categories → 'UAL'."""
    return 'UAL' if str(cat).upper().startswith('UAL') else str(cat)


# ── Excel data cache ──────────────────────────────────────────────────────────
_data_cache: dict = {'tasks': None, 'mtime': None, 'timestamp': None}
_ual_cache:  dict = {'rows': None, 'mtime': None, 'timestamp': None}

# ── SharePoint auto-download ─────────────────────────────────────────────────
_SP_GRAPH_TOKEN_FILE = SOC_AUTOMATION_DIR / '.graph_token_cache.json'
_SP_DRIVE_ID  = 'b!DoVu96jyeUi6KrTdnfiyOkD3l4uMELNPlWLzwTS0EvRl5FXkvP32Rr3npiSn6z7D'
_SP_ITEM_ID   = '01P5EGRYAB7M2ECPZAGBCZPOQMJ44PSEED'
_SP_CLIENT_ID = 'd3590ed6-52b3-4102-aeff-aad2292ab01c'
_SP_TENANT    = '46c98d88-e344-4ed4-8496-4ed7712e255d'
_SP_EXCEL_TTL = 300  # re-download at most every 5 minutes
_sp_dl_cache: dict = {'timestamp': None}


def _get_graph_token() -> str | None:
    """Return a valid Microsoft Graph access token, refreshing if needed."""
    if not _SP_GRAPH_TOKEN_FILE.exists():
        return None
    try:
        td = json.loads(_SP_GRAPH_TOKEN_FILE.read_text())
        # Refresh if expired or within 60s of expiry
        if datetime.now().timestamp() > td.get('expires_at', 0) - 60:
            r = _requests.post(
                f'https://login.microsoftonline.com/{_SP_TENANT}/oauth2/v2.0/token',
                data={
                    'client_id':     _SP_CLIENT_ID,
                    'scope':         'https://graph.microsoft.com/.default offline_access',
                    'refresh_token': td.get('refresh_token', ''),
                    'grant_type':    'refresh_token',
                },
                timeout=20,
            )
            if r.status_code != 200:
                logger.warning(f'Graph token refresh failed: {r.status_code}')
                return None
            nd = r.json()
            td['access_token'] = nd['access_token']
            td['expires_at']   = datetime.now().timestamp() + nd.get('expires_in', 3600)
            if 'refresh_token' in nd:
                td['refresh_token'] = nd['refresh_token']
            _SP_GRAPH_TOKEN_FILE.write_text(json.dumps(td))
        return td['access_token']
    except Exception as e:
        logger.warning(f'Graph token load/refresh error: {e}')
        return None


def _refresh_excel_from_sharepoint() -> bool:
    """Download JGS_SOC_Trend.xlsx from SharePoint and replace local copy.

    Returns True if download succeeded, False otherwise.
    Uses _SP_EXCEL_TTL to avoid hammering SharePoint on every request.
    """
    global _sp_dl_cache, _data_cache, _ual_cache
    now = datetime.now().timestamp()
    if (
        _sp_dl_cache['timestamp'] is not None
        and now - _sp_dl_cache['timestamp'] < _SP_EXCEL_TTL
    ):
        return True  # recently refreshed, skip

    token = _get_graph_token()
    if not token:
        return False

    try:
        excel_path = SOC_AUTOMATION_DIR / 'JGS_SOC_Trend.xlsx'
        r = _requests.get(
            f'https://graph.microsoft.com/v1.0/drives/{_SP_DRIVE_ID}/items/{_SP_ITEM_ID}/content',
            headers={'Authorization': f'Bearer {token}'},
            allow_redirects=True,
            timeout=60,
        )
        if r.status_code != 200:
            logger.warning(f'SharePoint Excel download failed: {r.status_code}')
            return False

        # Write atomically via temp file
        tmp = excel_path.with_suffix('.xlsx.sp_tmp')
        tmp.write_bytes(r.content)
        tmp.replace(excel_path)

        # Bust data caches so next read picks up fresh file
        _data_cache = {'tasks': None, 'mtime': None, 'timestamp': None}
        _ual_cache  = {'rows': None,  'mtime': None, 'timestamp': None}

        _sp_dl_cache['timestamp'] = now
        logger.info(f'SharePoint Excel refreshed ({len(r.content):,} bytes)')
        return True
    except Exception as e:
        logger.warning(f'SharePoint Excel refresh error: {e}')
        return False

# ── SQLite DB for UAL cell edits (shared across all users) ──────────────────
_SOC_DB = SOC_AUTOMATION_DIR / 'soc_edits.db'

@contextmanager
def _db_conn():
    conn = sqlite3.connect(str(_SOC_DB))
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()

def _init_soc_db():
    with _db_conn() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS ual_edits (
                row_id     TEXT NOT NULL,
                col_idx    INTEGER NOT NULL,
                value      TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL DEFAULT '',
                PRIMARY KEY (row_id, col_idx)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS task_comments (
                jira_key   TEXT PRIMARY KEY NOT NULL,
                value      TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL DEFAULT ''
            )
        ''')

_init_soc_db()


def _load_task_comments() -> dict:
    """Return {jira_key: comment_text} from DB."""
    with _db_conn() as conn:
        rows = conn.execute('SELECT jira_key, value FROM task_comments').fetchall()
    return {r['jira_key']: r['value'] for r in rows}


def _save_task_comment(jira_key: str, value: str):
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    with _db_conn() as conn:
        conn.execute(
            'INSERT INTO task_comments (jira_key, value, updated_at) VALUES (?,?,?)'
            ' ON CONFLICT(jira_key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at',
            (jira_key, value, now)
        )

def _load_ual_edits() -> dict:
    """Return {row_id: {col_idx: value}} from DB."""
    with _db_conn() as conn:
        rows = conn.execute('SELECT row_id, col_idx, value FROM ual_edits').fetchall()
    result: dict = {}
    for r in rows:
        result.setdefault(r['row_id'], {})[r['col_idx']] = r['value']
    return result

def _save_ual_edit(row_id: str, col_idx: int, value: str):
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    with _db_conn() as conn:
        conn.execute(
            'INSERT INTO ual_edits (row_id, col_idx, value, updated_at) VALUES (?,?,?,?)'
            ' ON CONFLICT(row_id, col_idx) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at',
            (row_id, col_idx, value, now)
        )

# ── Jira label status cache (keyed by issue key → set of series) ─────────────
_jira_label_cache: dict = {'status': None, 'timestamp': None}
_JIRA_LABEL_TTL = 300  # seconds — refresh every 5 minutes

# ── Jira blocked-by cache (keyed by VLK issue key → list of non-VLK blockers) ──
_blocked_by_cache: dict = {'data': None, 'timestamp': None}
_BLOCKED_BY_TTL  = 300  # seconds
# Categories where blockers come from Jira "is blocked by" links (non-VLK, open only)
_JIRA_BLOCKER_CATS = frozenset(['PSOC Boot/Load', 'Boot MUC/NUC', 'Boot GT'])
# Categories where blockers come directly from Excel col K
_EXCEL_BLOCKER_CATS = frozenset(['Boot in SRIOV Mode'])


def _get_jira_label_status() -> dict:
    """Return cached Jira label status or fetch fresh from Jira API.

    Returns { "VLK-12345": {"cc", "sim"}, ... }
    Falls back to empty dict on any error so the page still loads.
    """
    global _jira_label_cache
    now = datetime.now().timestamp()
    if (
        _jira_label_cache['status'] is not None
        and _jira_label_cache['timestamp'] is not None
        and now - _jira_label_cache['timestamp'] < _JIRA_LABEL_TTL
    ):
        return _jira_label_cache['status']

    if not JIRA_TOKEN_PATH.exists():
        return {}
    try:
        from fetch_jira_data import get_issue_label_status as _get_lbl
        pat = JIRA_TOKEN_PATH.read_text(encoding='utf-8-sig').strip()
        if not pat:
            return {}
        result = _get_lbl(pat)
        _jira_label_cache['status']    = result
        _jira_label_cache['timestamp'] = now
        logger.info(f"Jira label status refreshed: {len(result)} issues with labels")
        return result
    except Exception as e:
        logger.warning(f"Jira label status fetch failed: {e}")
        return _jira_label_cache.get('status') or {}


def _jira_get_no_ssl(pat: str, path: str, params: dict = None) -> dict:
    """Jira GET with SSL verification disabled (self-signed cert environment)."""
    url = 'https://jira.devtools.intel.com' + path
    if params:
        url += '?' + urllib.parse.urlencode(params)
    req = urllib.request.Request(url)
    req.add_header('Authorization', f'Bearer {pat}')
    req.add_header('Accept', 'application/json')
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    with urllib.request.urlopen(req, context=ctx, timeout=30) as r:
        return json.loads(r.read())


def _get_blocked_by_map(vlk_keys: list) -> dict:
    """Fetch 'is blocked by' Jira links for the given VLK issue keys.

    Returns {vlk_key: [blocker_key, ...]} where blocker_key is non-VLK
    and the blocker issue is not in a closed/rejected state.
    Cached for _BLOCKED_BY_TTL seconds.  Falls back to empty dict on error.
    """
    global _blocked_by_cache
    now = datetime.now().timestamp()
    if (
        _blocked_by_cache['data'] is not None
        and _blocked_by_cache['timestamp'] is not None
        and now - _blocked_by_cache['timestamp'] < _BLOCKED_BY_TTL
    ):
        return _blocked_by_cache['data']

    if not vlk_keys or not JIRA_TOKEN_PATH.exists():
        return {}

    try:
        pat = JIRA_TOKEN_PATH.read_text(encoding='utf-8-sig').strip()
        if not pat:
            return {}

        _CLOSED = {'closed', 'resolved', 'done', 'rejected', 'cancelled',
                   "won't fix", 'wontfix', 'invalid', 'duplicate'}

        result: dict = {k: [] for k in vlk_keys}
        start, page_size = 0, 50
        jql = 'issueKey in (' + ','.join(vlk_keys) + ')'
        while True:
            data = _jira_get_no_ssl(pat, '/rest/api/2/search', {
                'jql':        jql,
                'fields':     'issuelinks',
                'maxResults': page_size,
                'startAt':    start,
            })
            for issue in data.get('issues', []):
                key = issue['key']
                for lnk in issue.get('fields', {}).get('issuelinks', []):
                    # Inward issue = current issue "is blocked by" inwardIssue
                    if lnk.get('type', {}).get('inward', '').lower() != 'is blocked by':
                        continue
                    blocker = lnk.get('inwardIssue')
                    if not blocker:
                        continue
                    bkey = blocker['key']
                    if bkey.upper().startswith('VLK-'):
                        continue
                    bstatus = (blocker.get('fields', {}).get('status', {}).get('name', '') or '').lower()
                    bcat = (blocker.get('fields', {}).get('status', {}).get(
                            'statusCategory', {}).get('name', '') or '').lower()
                    if bstatus in _CLOSED or bcat in ('done', 'complete'):
                        continue
                    if key in result:
                        result[key].append(bkey)
            total   = data.get('total', 0)
            start  += page_size
            if start >= total:
                break

        _blocked_by_cache['data']      = result
        _blocked_by_cache['timestamp'] = now
        logger.info(f'Blocked-by map fetched for {len(vlk_keys)} VLK keys')
        return result

    except Exception as e:
        logger.warning(f'Blocked-by Jira fetch failed: {e}')
        return _blocked_by_cache.get('data') or {}


def invalidate_cache() -> None:
    """Force the next request to reload data from Excel and Jira."""
    global _data_cache, _jira_label_cache, _ual_cache, _blocked_by_cache, _sp_dl_cache
    _data_cache       = {'tasks': None, 'mtime': None, 'timestamp': None}
    _ual_cache        = {'rows': None, 'mtime': None, 'timestamp': None}
    _jira_label_cache = {'status': None, 'timestamp': None}
    _blocked_by_cache = {'data': None, 'timestamp': None}
    _sp_dl_cache      = {'timestamp': None}


# ── Excel reader ──────────────────────────────────────────────────────────────

def _ww_int(ww: str) -> int | None:
    """Convert a WW string like '26ww13' or 'ww13' to a sortable int."""
    m = _re.match(r'(?:(\d{2}))?[Ww][Ww](\d{1,2})', str(ww or '').strip())
    if not m:
        return None
    year = int(m.group(1)) if m.group(1) else 0
    return year * 100 + int(m.group(2))


def _load_date_history() -> dict:
    """Load date_change_history.json.

    Returns { jira_key: { field: [ {date, value}, ... ] } }
    Falls back to _read_old_dates() snapshot if history file is missing.
    """
    history_file = SOC_AUTOMATION_DIR / 'date_change_history.json'
    if history_file.exists():
        try:
            with open(history_file) as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Could not read date_change_history.json: {e}")

    # Fallback: build a single-entry history from backup Excel
    import io
    backup_file = SOC_AUTOMATION_DIR / 'JGS_SOC_Trend.xlsx.backup'
    if not backup_file.exists():
        return {}
    try:
        with open(backup_file, 'rb') as f:
            data = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(data, data_only=True)
        ws = wb['RTLFreeze']
        history: dict = {}
        for row in range(2, ws.max_row + 1):
            jira_key = ws.cell(row=row, column=12).value
            if not jira_key or str(jira_key).strip() in ('', '??'):
                continue
            jk = str(jira_key).strip()
            history[jk] = {
                'palladium_avail': [{'date': '', 'value': str(ws.cell(row=row, column=8).value  or '').strip()}],
                'code_trend':      [{'date': '', 'value': str(ws.cell(row=row, column=13).value or '').strip()}],
                'sim_trend':       [{'date': '', 'value': str(ws.cell(row=row, column=14).value or '').strip()}],
                'emu_trend':       [{'date': '', 'value': str(ws.cell(row=row, column=15).value or '').strip()}],
            }
        return history
    except Exception as e:
        logger.warning(f"Could not read backup Excel for date comparison: {e}")
        return {}


def _date_arrow_info(history_chain: list, current_value: str) -> dict:
    """Given a history chain and the current value, return rendering metadata.

    Returns:
      {
        'show':      bool,           # True if there is a change to display
        'old':       str,            # previous value
        'new':       str,            # current value
        'direction': 'in'|'out'|'', # pulled-in (green) or pushed-out (red)
        'tooltip':   str,            # full history chain as text
      }
    """
    if not history_chain:
        return {'show': False, 'old': '', 'new': current_value, 'direction': '', 'tooltip': ''}

    # Build full tooltip string (all recorded snapshots + current)
    parts = [f"{e['date']}: {e['value']}" if e.get('date') else e['value']
             for e in history_chain]
    # Append current if it differs from last in chain
    last_recorded = history_chain[-1]['value']
    if current_value and current_value.lower() != last_recorded.lower():
        from datetime import date
        parts.append(f"{date.today().isoformat()}: {current_value}")
        display_old = last_recorded
        display_new = current_value
    elif len(history_chain) >= 2:
        display_old = history_chain[-2]['value']
        display_new = last_recorded
    else:
        return {'show': False, 'old': '', 'new': current_value, 'direction': '', 'tooltip': ' → '.join(parts)}

    if display_old.lower() == display_new.lower():
        return {'show': False, 'old': '', 'new': current_value, 'direction': '', 'tooltip': ' → '.join(parts)}

    # Compute direction: earlier WW = pulled in (good=green), later = pushed out (bad=red)
    direction = ''
    old_int = _ww_int(display_old)
    new_int = _ww_int(display_new)
    if old_int is not None and new_int is not None:
        direction = 'in' if new_int < old_int else 'out'

    return {
        'show':      True,
        'old':       display_old,
        'new':       display_new,
        'direction': direction,
        'tooltip':   ' → '.join(parts),
    }


def get_rtl_freeze_data() -> list[dict]:
    """Read RTL Freeze data from Excel (cached, 5-min TTL)."""
    excel_file       = SOC_AUTOMATION_DIR / 'JGS_SOC_Trend.xlsx'
    tracking_db_file = SOC_AUTOMATION_DIR / 'task_tracking_database.json'

    try:
        now       = datetime.now().timestamp()
        file_mtime = excel_file.stat().st_mtime

        if (
            _data_cache['tasks'] is not None
            and _data_cache['mtime'] == file_mtime
            and _data_cache['timestamp'] is not None
            and now - _data_cache['timestamp'] < CACHE_TTL
        ):
            return _data_cache['tasks']

        wb_formatted = openpyxl.load_workbook(excel_file, data_only=False)
        wb_data      = openpyxl.load_workbook(excel_file, data_only=True)
        ws_fmt       = wb_formatted['RTLFreeze']
        ws_data      = wb_data['RTLFreeze']

        tracking_db: dict = {}
        if tracking_db_file.exists():
            with open(tracking_db_file) as f:
                tracking_db = json.load(f)

        # Fetch Jira label status to supplement Excel cell-colour checks
        jira_labels = _get_jira_label_status()

        tasks = []
        for row in range(2, ws_data.max_row + 1):
            jira_key        = ws_data.cell(row=row, column=12).value  # L
            category        = ws_data.cell(row=row, column=1).value   # A
            task_name       = ws_data.cell(row=row, column=2).value   # B
            owner           = ws_data.cell(row=row, column=3).value   # C
            palladium_avail = ws_data.cell(row=row, column=8).value   # H
            status          = ws_data.cell(row=row, column=10).value  # J
            blockers        = ws_data.cell(row=row, column=11).value  # K
            code_trend      = ws_data.cell(row=row, column=13).value  # M
            sim_trend       = ws_data.cell(row=row, column=14).value  # N
            emu_trend       = ws_data.cell(row=row, column=15).value  # O

            jira_key = str(jira_key or '').strip()
            if jira_key in ('', '??'):
                # Include rows without a JIRA if they have meaningful data
                if not (category or task_name):
                    continue
                jira_key = f'#ROW-{row}'

            if (not category or str(category).strip() == '') and (
                not task_name or str(task_name).strip() == ''
            ):
                continue

            def _has_valid_trend(v) -> bool:
                """Return True only when v is a real WW value, not N/A / blank / ??."""
                return bool(v and str(v).strip().lower() not in ('', 'n/a', '??', '—', 'none'))

            code_complete = is_cell_green(ws_fmt.cell(row=row, column=13)) and _has_valid_trend(code_trend)
            sim_complete  = is_cell_green(ws_fmt.cell(row=row, column=14)) and _has_valid_trend(sim_trend)
            emu_complete  = is_cell_green(ws_fmt.cell(row=row, column=15)) and _has_valid_trend(emu_trend)

            completions = tracking_db.get('task_completions', {})
            if _has_valid_trend(code_trend) and jira_key in completions.get('code_complete', {}): code_complete = True
            if _has_valid_trend(sim_trend)  and jira_key in completions.get('sim_verified',  {}): sim_complete  = True
            if _has_valid_trend(emu_trend)  and jira_key in completions.get('emu_verified',  {}): emu_complete  = True

            # Also check live Jira labels (Code_Complete / Coral_Tested / Palladium_Tested)
            jira_series = jira_labels.get(jira_key, set())
            if _has_valid_trend(code_trend) and 'cc'  in jira_series: code_complete = True
            if _has_valid_trend(sim_trend)  and 'sim' in jira_series: sim_complete  = True
            if _has_valid_trend(emu_trend)  and 'emu' in jira_series: emu_complete  = True

            cat_str = str(category).strip() if category else ''
            tasks.append({
                'jira_key':        jira_key,
                'category':        cat_str,
                'graph_category':  _graph_category(cat_str),
                'task_name':       str(task_name).strip()       if task_name       else '',
                'owner':           str(owner).strip()           if owner           else '',
                'palladium_avail': str(palladium_avail).strip() if palladium_avail else '',
                'status':          str(status).strip()          if status          else '',
                'blockers':        str(blockers).strip()        if blockers        else '',
                'code_trend':      str(code_trend).strip()      if code_trend      else '',
                'sim_trend':       str(sim_trend).strip()       if sim_trend       else '',
                'emu_trend':       str(emu_trend).strip()       if emu_trend       else '',
                'code_complete':   code_complete,
                'sim_complete':    sim_complete,
                'emu_complete':    emu_complete,
            })

        # Merge date history for arrow display
        date_history = _load_date_history()
        for task in tasks:
            hist = date_history.get(task['jira_key'], {})
            task['arr_pal']  = _date_arrow_info(hist.get('palladium_avail', []), task['palladium_avail'])
            task['arr_code'] = _date_arrow_info(hist.get('code_trend',      []), task['code_trend'])
            task['arr_sim']  = _date_arrow_info(hist.get('sim_trend',       []), task['sim_trend'])
            task['arr_emu']  = _date_arrow_info(hist.get('emu_trend',       []), task['emu_trend'])
            # Baseline (original) planned dates — first-ever recorded value;
            # used for the planned line on the graph so it stays fixed even
            # when dates get pushed out or pulled in.
            def _orig(chain, current):
                return chain[0]['value'] if chain and chain[0].get('value') else current
            task['code_trend_orig'] = _orig(hist.get('code_trend', []), task['code_trend'])
            task['sim_trend_orig']  = _orig(hist.get('sim_trend',  []), task['sim_trend'])
            task['emu_trend_orig']  = _orig(hist.get('emu_trend',  []), task['emu_trend'])

        _data_cache['tasks']     = tasks
        _data_cache['mtime']     = file_mtime
        _data_cache['timestamp'] = now
        return tasks

    except Exception as e:
        logger.error(f"Error reading RTL Freeze data: {e}")
        return []


def get_ual_schedule_data() -> list[dict]:
    """Read UAL_TI_SCHEDULE tab (cols A-I) from the Excel, merged with DB edits.

    Returns list of dicts:
      { 'row_id', 'headers'(only row 0), 'cells': [{value, is_green, col_idx}], 'jira_keys' }
    Falls back to empty list if the tab doesn't exist yet.
    """
    excel_file = SOC_AUTOMATION_DIR / 'JGS_SOC_Trend.xlsx'
    if not excel_file.exists():
        return []

    try:
        mtime = excel_file.stat().st_mtime
        now   = datetime.now().timestamp()
        if (
            _ual_cache['rows'] is not None
            and _ual_cache['mtime'] == mtime
            and _ual_cache['timestamp'] is not None
            and now - _ual_cache['timestamp'] < CACHE_TTL
        ):
            return _ual_cache['rows']

        wb_fmt  = openpyxl.load_workbook(excel_file, data_only=False)
        wb_data = openpyxl.load_workbook(excel_file, data_only=True)

        if 'UAL_TI_SCHEDULE' not in wb_data.sheetnames:
            _ual_cache.update({'rows': [], 'mtime': mtime, 'timestamp': now})
            return []

        ws_fmt  = wb_fmt['UAL_TI_SCHEDULE']
        ws_data = wb_data['UAL_TI_SCHEDULE']

        max_col = ws_data.max_column or 11  # read all columns as-is from Excel
        # Row 1 = headers
        headers = [str(ws_data.cell(row=1, column=c).value or '').strip() for c in range(1, max_col + 1)]

        rows: list[dict] = []
        for row_idx in range(2, ws_data.max_row + 1):
            cells = []
            has_data = False
            for col_idx in range(1, max_col + 1):
                cell_data = ws_data.cell(row=row_idx, column=col_idx)
                cell_fmt  = ws_fmt.cell(row=row_idx, column=col_idx)
                raw_val   = str(cell_data.value or '')
                if raw_val.strip() and raw_val.strip() not in ('None', 'nan'):
                    has_data = True
                cell_jiras    = _re.findall(r'[A-Z][A-Z0-9]+-\d+', raw_val)
                remaining_txt = _re.sub(r'[A-Z][A-Z0-9]+-\d+', '', raw_val).strip(' ,;/')
                cells.append({
                    'value':          raw_val,
                    'is_green':       is_cell_green(cell_fmt),
                    'col_idx':        col_idx,
                    'cell_jiras':     cell_jiras,
                    'remaining_text': remaining_txt,
                })
            if not has_data:
                continue

            row_id = f'ual_{row_idx}'
            # No DB overrides — display exactly what is in Excel

            # Extract Jira keys from all cells
            all_text = ' '.join(c['value'] for c in cells)
            jira_keys = _re.findall(r'[A-Z][A-Z0-9]+-\d+', all_text)

            rows.append({
                'row_id':    row_id,
                'cells':     cells,
                'jira_keys': list(dict.fromkeys(jira_keys)),  # deduped, ordered
            })

        _ual_cache.update({'rows': rows, 'mtime': mtime, 'timestamp': now})
        return rows

    except Exception as e:
        logger.error(f"Error reading UAL_TI_SCHEDULE: {e}")
        return []


def get_category_tasks() -> dict:
    """Return RTLFreeze tasks grouped by category (ordered by first appearance)."""
    tasks = get_rtl_freeze_data()
    grouped: dict = {}
    for t in tasks:
        cat = t['category']
        if cat not in grouped:
            grouped[cat] = []
        grouped[cat].append(t)
    return grouped


def get_ual_headers() -> list[str]:
    """Return column A-I header names from UAL_TI_SCHEDULE (or generic fallback)."""
    excel_file = SOC_AUTOMATION_DIR / 'JGS_SOC_Trend.xlsx'
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if 'UAL_TI_SCHEDULE' not in wb.sheetnames:
            return [f'Col {i}' for i in range(1, 10)]
        ws = wb['UAL_TI_SCHEDULE']
        max_col = ws.max_column or 11
        headers = [str(ws.cell(row=1, column=c).value or f'Col {c}').strip() for c in range(1, max_col + 1)]
        return headers
    except Exception:
        return [f'Col {i}' for i in range(1, 10)]


def get_long_pole_data() -> list[dict]:
    """Calculate the latest work week (long pole) per category (all UAL sub-cats kept separate)."""
    try:
        tasks      = get_rtl_freeze_data()
        categories: dict = {}
        for task in tasks:
            cat = task['category']
            if cat not in categories:
                categories[cat] = {'code_trends': [], 'sim_trends': [], 'emu_trends': []}
            if task['code_trend']: categories[cat]['code_trends'].append(task['code_trend'])
            if task['sim_trend']:  categories[cat]['sim_trends'].append(task['sim_trend'])
            if task['emu_trend']:  categories[cat]['emu_trends'].append(task['emu_trend'])

        def _latest(ww_list):
            dated = [(ww, parse_ww(ww)) for ww in ww_list]
            dated = [(ww, dt) for ww, dt in dated if dt is not None]
            return max(dated, key=lambda x: x[1])[0] if dated else None

        long_poles = []
        for cat in dict.fromkeys(t['category'] for t in tasks if t['category']):
            d = categories[cat]
            long_poles.append({
                'category':       cat,
                'code_long_pole': _latest(d['code_trends']) or 'N/A',
                'sim_long_pole':  _latest(d['sim_trends'])  or 'N/A',
                'emu_long_pole':  _latest(d['emu_trends'])  or 'N/A',
            })
        return long_poles

    except Exception as e:
        logger.error(f"Error calculating long pole data: {e}")
        return []


# ── Graph data helpers ────────────────────────────────────────────────────────

def _gww_int(ww) -> int | None:
    m = _re.match(r'(\d+)[Ww]{2}(\d+)', str(ww or ''))
    return int(m.group(1)) * 100 + int(m.group(2)) if m else None


def _gww_label(yy: int, wk: int) -> str:
    return f'{yy}WW{wk:02d}'


def _cum_trend(task_list: list, key: str, graph_labels: list, cur_int: int) -> list:
    counts: Counter = Counter()
    last_label = graph_labels[-1]
    for t in task_list:
        wi = _gww_int(t.get(key, ''))
        if wi and 2600 <= wi <= 2703:
            lbl = _gww_label(wi // 100, wi % 100)
            # cap beyond-range tasks to last bucket so they still count
            counts[lbl if lbl in graph_labels else last_label] += 1
    result, running = [], 0
    for lbl in graph_labels:
        running += counts.get(lbl, 0)
        result.append(running)
    return result


def _cum_done(task_list: list, trend_key: str, done_key: str,
              graph_labels: list, current_ww: str, cur_int: int) -> list:
    counts: Counter = Counter()
    for t in task_list:
        if t.get(done_key):
            wi = _gww_int(t.get(trend_key, ''))
            if wi and 2600 <= wi <= 2703:
                eff = min(wi, cur_int)
                counts[_gww_label(eff // 100, eff % 100)] += 1
    result, running, past = [], 0, True
    for lbl in graph_labels:
        running += counts.get(lbl, 0)
        result.append(running if past else None)
        if lbl == current_ww:
            past = False
    return result


def _long_pole(task_list: list) -> str | None:
    best = None
    for key in ('code_trend', 'sim_trend', 'emu_trend'):
        for t in task_list:
            wi = _gww_int(t.get(key, ''))
            if wi and (best is None or wi > best):
                best = wi
    return _gww_label(best // 100, best % 100) if best else None


def _parse_tw_as_ww_int(cell_value: str, first: bool = False) -> int | None:
    """Extract a TW\\d+ from a UAL effort cell and return as WW int (e.g. TW13 → 2613).

    first=True  → first TW (original planned date, before any pushes) — used for trend lines
    first=False → last  TW (most recent date) — used for done-bar placement
    """
    matches = _re.findall(r'TW(\d+)', cell_value or '', _re.IGNORECASE)
    if not matches:
        return None
    tw = matches[0] if first else matches[-1]
    return 2600 + int(tw)  # treat TW as WW in year 26


def _ual_cum_done(ual_rows: list, col_idx: int, graph_labels: list,
                  current_ww: str, cur_int: int) -> list:
    """Cumulative done bars for UAL: green cells in col_idx, WW derived from TW."""
    counts: Counter = Counter()
    for row in ual_rows:
        cell = next((c for c in row['cells'] if c['col_idx'] == col_idx), None)
        if cell and cell['is_green']:
            wi = _parse_tw_as_ww_int(cell['value'])
            if wi and 2600 <= wi <= 2703:
                eff = min(wi, cur_int)
                counts[_gww_label(eff // 100, eff % 100)] += 1
    result, running, past = [], 0, True
    for lbl in graph_labels:
        running += counts.get(lbl, 0)
        result.append(running if past else None)
        if lbl == current_ww:
            past = False
    return result


def _ual_cum_trend(ual_rows: list, col_idx: int, graph_labels: list, cur_int: int) -> list:
    """Cumulative planned trend for UAL: all cells in col_idx, first TW = original planned date."""
    counts: Counter = Counter()
    last_label = graph_labels[-1]
    for row in ual_rows:
        cell = next((c for c in row['cells'] if c['col_idx'] == col_idx), None)
        if cell:
            wi = _parse_tw_as_ww_int(cell['value'], first=True)  # original date before pushes
            if wi and 2600 <= wi <= 2703:
                lbl = _gww_label(wi // 100, wi % 100)
                # cap beyond-range tasks to last bucket so they still count
                counts[lbl if lbl in graph_labels else last_label] += 1
    result, running = [], 0
    for lbl in graph_labels:
        running += counts.get(lbl, 0)
        result.append(running)
    return result


def _ual_long_pole(ual_rows: list) -> str | None:
    """Latest TW (as WW) across all effort columns in UAL_TI_SCHEDULE."""
    best = None
    for col_idx in (4, 6, 8):
        for row in ual_rows:
            cell = next((c for c in row['cells'] if c['col_idx'] == col_idx), None)
            if cell:
                wi = _parse_tw_as_ww_int(cell['value'])
                if wi and (best is None or wi > best):
                    best = wi
    return _gww_label(best // 100, best % 100) if best else None


# ── Route ─────────────────────────────────────────────────────────────────────

@bp.route('/jgs-soc-updates')
def jgs_soc_updates():
    # Always try to pull the latest Excel from SharePoint first (cached 5 min)
    _refresh_excel_from_sharepoint()

    tasks         = get_rtl_freeze_data()
    long_pole_data = get_long_pole_data()
    current_ww, cur_int, current_ww_date = current_ww_info()

    for task in tasks:
        task['blocker_jiras'] = extract_jira_keys(task['blockers']) if task['blockers'] else []

    # Pending tasks
    pending_tasks = []
    for task in tasks:
        for label, trend_key, complete_key in [
            ('Code Complete',   'code_trend', 'code_complete'),
            ('Coral Tested',    'sim_trend',  'sim_complete'),
            ('Palladium Tested','emu_trend',  'emu_complete'),
        ]:
            trend_ww  = task[trend_key]
            if not trend_ww or task[complete_key]:
                continue
            trend_date = parse_ww(trend_ww)
            if trend_date and trend_date <= current_ww_date:
                pending_tasks.append({
                    'jira_key':      task['jira_key'],
                    'task_name':     task['task_name'],
                    'category':      task['category'],
                    'graph_category': task['graph_category'],
                    'type':          label,
                    'expected_ww':   trend_ww,
                    'is_overdue':    trend_date < current_ww_date,
                })

    blocker_tasks = [t for t in tasks if t['blockers'] and t['blockers'] != 'None']

    # Graph data
    graph_ww_labels = [_gww_label(26, wk) for wk in range(1, 27)]

    soc_graph_data = {
        'labels':    graph_ww_labels,
        'cc_trend':  _cum_trend(tasks, 'code_trend_orig', graph_ww_labels, cur_int),
        'sim_trend': _cum_trend(tasks, 'sim_trend_orig',  graph_ww_labels, cur_int),
        'emu_trend': _cum_trend(tasks, 'emu_trend_orig',  graph_ww_labels, cur_int),
    }

    # Build graph/long-pole data grouped by graph_category (UAL* consolidated)
    _graph_cats = list(dict.fromkeys(t['graph_category'] for t in tasks if t['graph_category']))
    _cat_data: dict = {}
    for _gcat in _graph_cats:
        _ct = [t for t in tasks if t['graph_category'] == _gcat]
        _cat_data[_gcat] = {
            'cc_trend':     _cum_trend(_ct, 'code_trend_orig', graph_ww_labels, cur_int),
            'sim_trend':    _cum_trend(_ct, 'sim_trend_orig',  graph_ww_labels, cur_int),
            'emu_trend':    _cum_trend(_ct, 'emu_trend_orig',  graph_ww_labels, cur_int),
            'cc_done':      _cum_done(_ct, 'code_trend', 'code_complete', graph_ww_labels, current_ww, cur_int),
            'sim_done':     _cum_done(_ct, 'sim_trend',  'sim_complete',  graph_ww_labels, current_ww, cur_int),
            'emu_done':     _cum_done(_ct, 'emu_trend',  'emu_complete',  graph_ww_labels, current_ww, cur_int),
            'long_pole_ww': _long_pole(_ct),
        }
    soc_category_data = {'categories': _graph_cats, 'data': _cat_data}

    # ── UAL from UAL_TI_SCHEDULE: override bars (green=done) + add UAL trend ──
    # UAL is not in RTLFreeze, so its tasks are absent from soc_graph_data totals.
    # We add the UAL trend contribution here so the "All" view line is complete.
    _ual_rows = get_ual_schedule_data()
    if _ual_rows:
        _ual_cc_trend  = _ual_cum_trend(_ual_rows, 4, graph_ww_labels, cur_int)
        _ual_sim_trend = _ual_cum_trend(_ual_rows, 6, graph_ww_labels, cur_int)
        _ual_emu_trend = _ual_cum_trend(_ual_rows, 8, graph_ww_labels, cur_int)

        # Add UAL into the "All" trend totals
        soc_graph_data['cc_trend']  = [a + b for a, b in zip(soc_graph_data['cc_trend'],  _ual_cc_trend)]
        soc_graph_data['sim_trend'] = [a + b for a, b in zip(soc_graph_data['sim_trend'], _ual_sim_trend)]
        soc_graph_data['emu_trend'] = [a + b for a, b in zip(soc_graph_data['emu_trend'], _ual_emu_trend)]

        # Add/replace UAL category entry in per-category data
        if 'UAL' not in _cat_data:
            _graph_cats.append('UAL')
        _cat_data['UAL'] = {
            'cc_trend':     _ual_cc_trend,
            'sim_trend':    _ual_sim_trend,
            'emu_trend':    _ual_emu_trend,
            'cc_done':      _ual_cum_done(_ual_rows, 4, graph_ww_labels, current_ww, cur_int),
            'sim_done':     _ual_cum_done(_ual_rows, 6, graph_ww_labels, current_ww, cur_int),
            'emu_done':     _ual_cum_done(_ual_rows, 8, graph_ww_labels, current_ww, cur_int),
            'long_pole_ww': _ual_long_pole(_ual_rows),
        }

    # Category blocker JIRAs on long-pole rows (keyed by category, not graph_category)
    _cat_blockers: dict = {}
    for task in tasks:
        cat = task['category']
        if cat not in _cat_blockers:
            _cat_blockers[cat] = set()
        _cat_blockers[cat].update(task.get('blocker_jiras', []))
    for item in long_pole_data:
        item['category_blocker_jiras'] = sorted(_cat_blockers.get(item['category'], []))

    # Platform stats (by graph_category)
    platform_stats = [
        {
            'group':      _gcat,
            'name':       _gcat,
            'fc_target':  sum(1 for t in [tt for tt in tasks if tt['graph_category'] == _gcat] if t['code_trend']),
            'sim_target': sum(1 for t in [tt for tt in tasks if tt['graph_category'] == _gcat] if t['sim_trend']),
            'emu_target': sum(1 for t in [tt for tt in tasks if tt['graph_category'] == _gcat] if t['emu_trend']),
            'fc_done':    sum(1 for t in [tt for tt in tasks if tt['graph_category'] == _gcat] if t['code_complete']),
            'sim_done':   sum(1 for t in [tt for tt in tasks if tt['graph_category'] == _gcat] if t['sim_complete']),
            'emu_done':   sum(1 for t in [tt for tt in tasks if tt['graph_category'] == _gcat] if t['emu_complete']),
        }
        for _gcat in _graph_cats
    ]

    category_tasks = get_category_tasks()

    # ── Enrich category_tasks with blocked_by_jiras per task ──
    # PSOC/Boot cats: Jira 'is blocked by' links (non-VLK, open issues only)
    _jira_vlk_keys = [
        t['jira_key']
        for cat, ct in category_tasks.items() if cat in _JIRA_BLOCKER_CATS
        for t in ct
        if t['jira_key'] and not t['jira_key'].startswith('#') and t['jira_key'].upper().startswith('VLK-')
    ]
    _blocked_by = _get_blocked_by_map(_jira_vlk_keys) if _jira_vlk_keys else {}

    for cat, ct in category_tasks.items():
        for t in ct:
            if cat in _JIRA_BLOCKER_CATS:
                t['blocked_by_jiras'] = _blocked_by.get(t['jira_key'], [])
            elif cat in _EXCEL_BLOCKER_CATS:
                t['blocked_by_jiras'] = extract_jira_keys(t.get('blockers', '') or '')
            else:
                t['blocked_by_jiras'] = []

    return render_template(
        'jgs_soc_updates.html',
        tasks=tasks,
        pending_tasks=pending_tasks,
        blocker_tasks=blocker_tasks,
        long_pole_data=long_pole_data,
        current_ww=current_ww,
        soc_graph_data=soc_graph_data,
        soc_category_data=soc_category_data,
        platform_stats=platform_stats,
        category_tasks=category_tasks,
        ual_rows=get_ual_schedule_data(),
        ual_headers=get_ual_headers(),
        task_comments=_load_task_comments(),
    )


@bp.route('/jgs-soc-updates/save-task-comment', methods=['POST'])
def jgs_soc_save_task_comment():
    """Persist a task comment edit to SQLite (shared across all users)."""
    try:
        data     = request.get_json(force=True)
        jira_key = str(data.get('jira_key', '')).strip()
        value    = str(data.get('value', '')).strip()
        if not jira_key:
            return jsonify({'ok': False, 'error': 'invalid params'}), 400
        _save_task_comment(jira_key, value)
        return jsonify({'ok': True})
    except Exception as e:
        logger.error(f"save-task-comment: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/jgs-soc-updates/save-ual-cell', methods=['POST'])
def jgs_soc_save_ual_cell():
    """Persist a UAL_TI_SCHEDULE cell edit to SQLite (shared across all users)."""
    try:
        data    = request.get_json(force=True)
        row_id  = str(data.get('row_id', '')).strip()
        col_idx = int(data.get('col_idx', 0))
        value   = str(data.get('value', '')).strip()
        if not row_id or not (1 <= col_idx <= 9):
            return jsonify({'ok': False, 'error': 'invalid params'}), 400
        _save_ual_edit(row_id, col_idx, value)
        # Bust UAL cache so next page load reflects edit
        global _ual_cache
        _ual_cache = {'rows': None, 'mtime': None, 'timestamp': None}
        return jsonify({'ok': True})
    except Exception as e:
        logger.error(f"save-ual-cell: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500

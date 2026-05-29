"""
JGS 4+2 Dashboard  —  /jgs-4plus2

Reads the JGS 4+2 tracking spreadsheet from Rodrigo Vivi's personal SharePoint
(OneDrive), parses Sheet1 to extract HSD + Jira + Owner rows, fetches HSD titles
via the HSD REST API (Kerberos), then renders an editable shared dashboard.

Owner → column mapping
    Rodrigo / Vivi  → xeKMD
    Erez            → UALkmd
    Mrozek          → UMD-L0
    Santosh         → Sysman

Routes
    GET  /jgs-4plus2                   Dashboard page
    POST /jgs-4plus2/save              Save ETA or Comments for one HSD row
    POST /jgs-4plus2/refresh-excel     Force re-download of the SharePoint Excel
    GET  /jgs-4plus2/api/data          JSON of all current edits (for polling)

Storage
    ETA + Comments are persisted in  Projects/JGS_4+2/jgs4plus2.db  (SQLite, WAL)
    so every user sees the same data on every page view / 60-second poll.
"""

import base64
import json
import logging
import re
import sqlite3
import time
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import requests
import urllib3
from flask import Blueprint, render_template, jsonify, request
from requests_kerberos import HTTPKerberosAuth, OPTIONAL

urllib3.disable_warnings()
logger = logging.getLogger(__name__)

bp = Blueprint('jgs_4plus2', __name__)

# ── Paths ─────────────────────────────────────────────────────────────────────
_ROUTES_DIR  = Path(__file__).parent           # data_centre_landing/routes/
_APP_DIR     = _ROUTES_DIR.parent              # data_centre_landing/
_PROJ_DIR    = _APP_DIR.parent                 # Projects/
_JGS4P2_DIR  = _PROJ_DIR / 'JGS_4+2'          # Projects/JGS_4+2/
_JGS4P2_DIR.mkdir(parents=True, exist_ok=True)

_EXCEL_PATH  = _JGS4P2_DIR / '_cache.xlsx'
_EXCEL_TS    = _JGS4P2_DIR / '_cache_ts.txt'
_DB_PATH     = _JGS4P2_DIR / 'jgs4plus2.db'

# ── SharePoint personal OneDrive share ────────────────────────────────────────
_SHARE_URL = (
    'https://intel-my.sharepoint.com/:x:/p/rodrigo_vivi/'
    'IQAaFUEocrEVQ7IEktmK5pjsAYKecGahv3hhu65Y1J1wr_s?e=py81bb'
)
_EXCEL_TTL = 3600   # seconds between automatic re-downloads

# Graph API token (shared with soc_automation)
_SOC_AUTO_DIR   = _PROJ_DIR / 'soc_automation'
_TOKEN_CACHE    = _SOC_AUTO_DIR / '.graph_token_cache.json'
_JIRA_TOKEN_PATH = _SOC_AUTO_DIR / '.jira_token'

# ── External URLs ─────────────────────────────────────────────────────────────
_HSD_REST  = 'https://hsdes.intel.com/rest/article/'
_HSD_APP   = 'https://hsdes.intel.com/appstore/article/#/'
_JIRA_BASE = 'https://jira.devtools.intel.com/browse/'
_HSD_TTL   = 86400   # cache HSD titles for 24 h

# ── In-memory caches ──────────────────────────────────────────────────────────
_excel_cache: dict = {'jgs': None, 'other': None, 'ts': 0.0, 'error': ''}
_hsd_cache:    dict = {}   # hsd_id → title
_hsd_cache_ts: dict = {}   # hsd_id → float timestamp


# ═══════════════════════════════════════════════════════════════════════════════
# SQLite helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _db_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(_DB_PATH), timeout=10)
    conn.execute('PRAGMA journal_mode=WAL')
    return conn


_JIRA_COLS  = ('jira_xekmd', 'jira_ualkmd', 'jira_umdl0', 'jira_sysman', 'jira_umd_sysman')
_VALID_FIELDS = {'eta', 'comments', 'ext_deps', 'jira_xekmd', 'jira_ualkmd', 'jira_umd_sysman'}


def _init_db() -> None:
    with _db_conn() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS edits (
                hsd_id          TEXT PRIMARY KEY,
                eta             TEXT NOT NULL DEFAULT '',
                comments        TEXT NOT NULL DEFAULT '',
                updated_at      TEXT NOT NULL DEFAULT '',
                jira_xekmd      TEXT NOT NULL DEFAULT '',
                jira_ualkmd     TEXT NOT NULL DEFAULT '',
                jira_umdl0      TEXT NOT NULL DEFAULT '',
                jira_sysman     TEXT NOT NULL DEFAULT '',
                ext_deps        TEXT NOT NULL DEFAULT '',
                jira_umd_sysman TEXT NOT NULL DEFAULT ''
            )
        ''')
        # Migrate existing tables
        for col in (*_JIRA_COLS, 'ext_deps'):
            try:
                conn.execute(f'ALTER TABLE edits ADD COLUMN {col} TEXT NOT NULL DEFAULT ""')
            except Exception:
                pass   # column already exists


def _load_edits() -> dict:
    """Return {hsd_id: {eta, comments, updated_at, jira_xekmd, jira_ualkmd, jira_umd_sysman, ext_deps}}."""
    _init_db()
    with _db_conn() as conn:
        rows = conn.execute(
            'SELECT hsd_id, eta, comments, updated_at,'
            ' jira_xekmd, jira_ualkmd, jira_umdl0, jira_sysman,'
            ' ext_deps, jira_umd_sysman FROM edits'
        ).fetchall()
    result = {}
    for r in rows:
        # Migrate: if new merged column is empty, seed it from legacy umdl0+sysman
        umd_sysman = r[9] or ''
        if not umd_sysman:
            legacy = ' '.join(filter(None, [r[6], r[7]]))
            umd_sysman = legacy
        result[r[0]] = {
            'eta': r[1], 'comments': r[2], 'updated_at': r[3],
            'jira_xekmd': r[4], 'jira_ualkmd': r[5],
            'jira_umd_sysman': umd_sysman,
            'ext_deps': r[8],
        }
    return result


def _save_edit(hsd_id: str, field: str, value: str) -> None:
    if field not in _VALID_FIELDS:
        raise ValueError(f'Invalid field: {field!r}')
    _init_db()
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    _DEFAULTS: dict = {f: '' for f in _VALID_FIELDS}
    with _db_conn() as conn:
        existing = conn.execute(
            'SELECT eta, comments, jira_xekmd, jira_ualkmd, jira_umd_sysman, ext_deps'
            ' FROM edits WHERE hsd_id = ?', (hsd_id,)
        ).fetchone()
        if existing:
            d = dict(zip(('eta', 'comments', 'jira_xekmd', 'jira_ualkmd',
                          'jira_umd_sysman', 'ext_deps'), existing))
            d[field] = value
            conn.execute(
                'UPDATE edits SET eta=?, comments=?, updated_at=?,'
                ' jira_xekmd=?, jira_ualkmd=?, jira_umd_sysman=?, ext_deps=?'
                ' WHERE hsd_id=?',
                (d['eta'], d['comments'], now,
                 d['jira_xekmd'], d['jira_ualkmd'], d['jira_umd_sysman'],
                 d['ext_deps'], hsd_id)
            )
        else:
            d = {**_DEFAULTS, field: value}
            conn.execute(
                'INSERT INTO edits'
                ' (hsd_id, eta, comments, updated_at, jira_xekmd, jira_ualkmd, jira_umd_sysman, ext_deps)'
                ' VALUES (?,?,?,?,?,?,?,?)',
                (hsd_id, d['eta'], d['comments'], now,
                 d['jira_xekmd'], d['jira_ualkmd'], d['jira_umd_sysman'],
                 d['ext_deps'])
            )


# ═══════════════════════════════════════════════════════════════════════════════
# SharePoint / Graph API helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _refresh_graph_token() -> bool:
    """Use the stored refresh_token to obtain a new access token. Returns True on success."""
    try:
        data = json.loads(_TOKEN_CACHE.read_text())
        refresh_tok = data.get('refresh_token')
        if not refresh_tok:
            logger.warning('No refresh_token in cache – cannot auto-refresh')
            return False

        logger.info('Graph token expired – attempting automatic refresh…')
        resp = requests.post(
            'https://login.microsoftonline.com/46c98d88-e344-4ed4-8496-4ed7712e255d/oauth2/v2.0/token',
            data={
                'client_id':     'd3590ed6-52b3-4102-aeff-aad2292ab01c',
                'scope':         'https://graph.microsoft.com/.default offline_access',
                'refresh_token': refresh_tok,
                'grant_type':    'refresh_token',
            },
            timeout=20,
        )
        if resp.status_code != 200:
            logger.warning('Auto-refresh failed: HTTP %s – %s', resp.status_code, resp.text[:200])
            return False

        new = resp.json()
        expires_at = (datetime.now() + timedelta(seconds=new.get('expires_in', 3600))).timestamp()
        _TOKEN_CACHE.write_text(json.dumps({
            'access_token':  new['access_token'],
            'refresh_token': new.get('refresh_token', refresh_tok),
            'expires_at':    expires_at,
        }, indent=2))
        logger.info('Graph token auto-refreshed successfully (expires %s)',
                    datetime.fromtimestamp(expires_at))
        return True
    except Exception as exc:
        logger.warning('Error during auto-refresh: %s', exc)
        return False


def _get_graph_token() -> str | None:
    if not _TOKEN_CACHE.exists():
        return None
    try:
        data = json.loads(_TOKEN_CACHE.read_text())
        if datetime.now().timestamp() > data.get('expires_at', 0):
            if not _refresh_graph_token():
                return None
            data = json.loads(_TOKEN_CACHE.read_text())
        return data.get('access_token')
    except Exception as exc:
        logger.warning('Cannot load Graph token: %s', exc)
        return None


def _encode_sharing_url(url: str) -> str:
    """Encode a sharing URL for the Graph API /shares/{encodedUrl}/driveItem endpoint."""
    return 'u!' + base64.urlsafe_b64encode(url.encode('utf-8')).decode('ascii').rstrip('=')


def _download_excel() -> tuple[bool, str]:
    """Download the SharePoint Excel and cache it locally.
    Returns (success, error_message).
    """
    token = _get_graph_token()
    if token:
        encoded  = _encode_sharing_url(_SHARE_URL)
        graph_url = f'https://graph.microsoft.com/v1.0/shares/{encoded}/driveItem/content'
        headers   = {'Authorization': f'Bearer {token}'}
        try:
            resp = requests.get(graph_url, headers=headers, timeout=30, allow_redirects=True)
            if resp.status_code == 200 and resp.content:
                _EXCEL_PATH.write_bytes(resp.content)
                _EXCEL_TS.write_text(str(time.time()))
                logger.info('JGS 4+2 Excel downloaded via Graph API (%d bytes)', len(resp.content))
                return True, ''
            logger.warning('Graph API returned HTTP %s for Excel download', resp.status_code)
        except Exception as exc:
            logger.warning('Graph API download failed: %s', exc)

    # Fallback: try a direct GET on the sharing URL (works for publicly shared links)
    try:
        resp = requests.get(_SHARE_URL, timeout=30, allow_redirects=True)
        # Verify it's actually an xlsx (PK magic bytes)
        if resp.status_code == 200 and resp.content[:4] == b'PK\x03\x04':
            _EXCEL_PATH.write_bytes(resp.content)
            _EXCEL_TS.write_text(str(time.time()))
            logger.info('JGS 4+2 Excel downloaded directly (%d bytes)', len(resp.content))
            return True, ''
        return False, (
            f'Direct download returned HTTP {resp.status_code}. '
            'Ensure the SharePoint file is shared and a valid Graph API token exists '
            'in soc_automation/.graph_token_cache.json.'
        )
    except Exception as exc:
        return False, f'Download error: {exc}'


def _excel_needs_refresh() -> bool:
    if not _EXCEL_PATH.exists():
        return True
    if not _EXCEL_TS.exists():
        return True
    try:
        return time.time() - float(_EXCEL_TS.read_text().strip()) > _EXCEL_TTL
    except Exception:
        return True


# ═══════════════════════════════════════════════════════════════════════════════
# Excel parsing
# ═══════════════════════════════════════════════════════════════════════════════

def _owner_to_col(owner: str) -> str:
    """Map an Eng Owner name (column H) to its dashboard column key."""
    o = (owner or '').lower().strip()
    # UAL KMD: Farah, Menny/Meny, Guy
    if any(n in o for n in ('farah', 'menny', 'meny', 'guy')):
        return 'ualkmd'
    # UMD / Sysman (merged): Ravi, Piotr
    if any(n in o for n in ('ravi', 'piotr')):
        return 'umd_sysman'
    # KMD: Saurabh, Anshuman, Alex, Stuart, James
    if any(n in o for n in ('saurabh', 'anshuman', 'alex', 'stuart', 'james')):
        return 'xekmd'
    return 'other'


def _parse_excel() -> tuple[list[dict], list[dict], str]:
    """Parse Sheet1 from the cached Excel file.

    Returns (jgs_groups, other_groups, error_message).
    Rows before the 'Other new HSDs' section header → jgs_groups.
    Rows from that header onwards → other_groups.
    Each group dict: {hsd_id, xekmd:[jira,...], ualkmd:[...], umdl0:[...], sysman:[...], other:[...]}
    """
    if not _EXCEL_PATH.exists():
        return [], [], 'Excel not yet downloaded'

    try:
        wb = openpyxl.load_workbook(str(_EXCEL_PATH), data_only=True)
    except Exception as exc:
        return [], [], f'Cannot open Excel: {exc}'

    # Find Sheet1 (try several spellings)
    ws = None
    for name in ('Sheet1', 'sheet1', 'Sheet 1', 'DATA', 'Data'):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]
        logger.info('JGS 4+2: "Sheet1" not found, using first sheet: %s', ws.title)

    # ── Detect header columns from row 1 ──────────────────────────────────────
    hsd_col = jira_col = owner_col = None
    for cell in ws[1]:
        v  = str(cell.value or '').strip().lower()
        ci = cell.column - 1
        if not v:
            continue
        if hsd_col is None and any(kw in v for kw in ('hsd', 'article', 'bug id', 'id')):
            hsd_col = ci
        elif jira_col is None and any(kw in v for kw in ('jira', 'ticket', 'vlk', 'key', 'issue')):
            jira_col = ci
        # Prefer 'Eng Owner' over 'Arch Owner' — match 'eng' + 'owner' together
        elif 'eng' in v and 'owner' in v:
            owner_col = ci
        elif owner_col is None and any(kw in v for kw in ('owner', 'assignee', 'responsible', 'dri')):
            owner_col = ci

    if hsd_col is None:
        hsd_col = 0   # default to column A

    # ── Read rows, splitting on the "Other new HSDs" section header ───────────
    raw_jgs:   list[dict] = []
    raw_other: list[dict] = []
    in_other_section = False
    last_hsd = ''    # carry-forward for merged cells in column A

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        hsd_raw = str(row[hsd_col] if hsd_col < len(row) else None or '').strip()
        if hsd_raw == 'None':
            hsd_raw = ''

        # Detect section separator: non-numeric text in HSD column that mentions
        # "other" — this marks the start of the second table.
        if hsd_raw and not re.match(r'^\d{7,}$', hsd_raw):
            if 'other' in hsd_raw.lower():
                in_other_section = True
                last_hsd = ''
            # Either way, this row is a section header — skip it
            last_hsd = ''
            continue

        if re.match(r'^\d{7,}$', hsd_raw):
            last_hsd = hsd_raw
        elif not hsd_raw:
            if not last_hsd:
                continue
            hsd_raw = last_hsd
        else:
            last_hsd = ''
            continue

        # Jira ID(s)
        jira_raw = ''
        if jira_col is not None and jira_col < len(row):
            jira_raw = str(row[jira_col] or '').strip()
        else:
            for idx, cell_val in enumerate(row):
                if idx == hsd_col:
                    continue
                v_str = str(cell_val or '')
                if re.search(r'[A-Za-z]+-\d+', v_str):
                    jira_raw = v_str
                    break

        jira_ids = [j.upper() for j in re.findall(r'[A-Za-z]+-\d+', jira_raw)]
        if not jira_ids and jira_raw and jira_raw not in ('', 'None', 'nan'):
            jira_ids = [jira_raw.strip()]

        # Owner
        owner_raw = ''
        if owner_col is not None and owner_col < len(row):
            owner_raw = str(row[owner_col] or '').strip()

        entry = {
            'hsd_id':  hsd_raw,
            'jira_ids': jira_ids,
            'owner':   owner_raw,
            'col':     _owner_to_col(owner_raw),
        }
        (raw_other if in_other_section else raw_jgs).append(entry)

    def _build_groups(raw_rows: list[dict]) -> list[dict]:
        groups: dict  = {}
        ordered: list = []
        for r in raw_rows:
            hid = r['hsd_id']
            if hid not in groups:
                groups[hid] = {
                    'hsd_id':     hid,
                    'xekmd':      [],
                    'ualkmd':     [],
                    'umd_sysman': [],
                    'other':      [],
                }
                ordered.append(hid)
            col = r['col']
            if col not in groups[hid]:
                col = 'other'
            for jid in r['jira_ids']:
                if jid not in groups[hid][col]:
                    groups[hid][col].append(jid)
        return [groups[hid] for hid in ordered]

    return _build_groups(raw_jgs), _build_groups(raw_other), ''


# ═══════════════════════════════════════════════════════════════════════════════
# HSD title fetch  +  Jira effort fetch
# ═══════════════════════════════════════════════════════════════════════════════

# Jira effort in-memory cache
_jira_effort_cache:    dict = {}   # jira_key → int seconds
_jira_effort_cache_ts: dict = {}   # jira_key → float timestamp
_JIRA_EFFORT_TTL = 14400           # 4 hours
_JIRA_REST = 'https://jira.devtools.intel.com/rest/api/2/issue/'


def _kerberos_session() -> requests.Session:
    s = requests.Session()
    s.auth   = HTTPKerberosAuth(mutual_authentication=OPTIONAL)
    s.verify = False
    return s


def _jira_session() -> requests.Session:
    """Return a requests.Session authenticated with the Jira PAT Bearer token."""
    s = requests.Session()
    s.verify = False
    try:
        token = _JIRA_TOKEN_PATH.read_text().strip()
        s.headers.update({'Authorization': f'Bearer {token}',
                          'Content-Type': 'application/json'})
    except Exception as exc:
        logger.warning('Cannot load Jira token from %s: %s', _JIRA_TOKEN_PATH, exc)
    return s


def _fetch_hsd_titles(hsd_ids: list[str]) -> dict:
    """Return {hsd_id: title_string}.  Uses an in-memory 24 h cache."""
    now      = time.time()
    to_fetch = [
        hid for hid in hsd_ids
        if hid not in _hsd_cache or now - _hsd_cache_ts.get(hid, 0) > _HSD_TTL
    ]
    if not to_fetch:
        return {hid: _hsd_cache.get(hid, '') for hid in hsd_ids}

    # Sequential fetch — Kerberos session is not thread-safe
    session = _kerberos_session()
    for hid in to_fetch:
        url = f'{_HSD_REST}{hid}'
        try:
            resp = session.get(url, timeout=12)
            if resp.status_code == 200:
                data  = resp.json()
                items = data.get('data', []) if isinstance(data, dict) else []
                if items and isinstance(items, list):
                    title = str(items[0].get('title', '')).strip()
                else:
                    title = str(data.get('title', '')).strip()
                _hsd_cache[hid]    = title
                _hsd_cache_ts[hid] = time.time()
            else:
                logger.warning('HSD %s returned HTTP %s', hid, resp.status_code)
                _hsd_cache[hid]    = ''
                _hsd_cache_ts[hid] = time.time()
        except Exception as exc:
            logger.debug('HSD title fetch %s: %s', hid, exc)
            _hsd_cache[hid]    = ''
            _hsd_cache_ts[hid] = time.time()

    return {hid: _hsd_cache.get(hid, '') for hid in hsd_ids}


def _format_effort_seconds(seconds: int) -> str:
    """Convert Jira seconds → total hours string (e.g. '80h')."""
    if not seconds or seconds <= 0:
        return '—'
    hours = round(seconds / 3600)
    return f'{hours}h' if hours > 0 else '<1h'


_JIRA_AGILE = 'https://jira.devtools.intel.com/rest/agile/1.0/epic/'


def _jira_child_keys(jira_key: str, fields: dict, session: requests.Session) -> list[str]:
    """Return child issue keys via explicit parent-child relationships only:
    1. subtasks  (standard Jira sub-tasks)
    2. issuelinks with type 'Child to Parent relations' (inward = child of this)
    Agile epic membership is handled separately in _jira_effort_seconds.
    """
    children: list[str] = []

    # 1. subtasks
    for sub in fields.get('subtasks', []):
        k = sub.get('key')
        if k:
            children.append(k)

    # 2. issuelinks — "Child to Parent relations" only, inward issue is the child
    for lnk in fields.get('issuelinks', []):
        if lnk.get('type', {}).get('name') == 'Child to Parent relations':
            inward = lnk.get('inwardIssue', {})
            k = inward.get('key') if inward else None
            if k:
                children.append(k)

    return children


def _jira_effort_seconds(jira_key: str, session: requests.Session,
                          depth: int = 0, visited: set | None = None) -> int:
    """Recursively fetch total original-estimate seconds for a Jira issue.

    Priority:
      1. timeoriginalestimate / aggregatetimeoriginalestimate on the issue itself
      2. Sum of children via subtasks, 'Child to Parent relations' issuelinks,
         or Agile epic membership (recurses up to depth 5).
    Results are cached for _JIRA_EFFORT_TTL seconds.
    """
    if visited is None:
        visited = set()
    if depth > 5 or jira_key in visited:
        return 0
    visited.add(jira_key)

    now = time.time()
    if jira_key in _jira_effort_cache and \
            now - _jira_effort_cache_ts.get(jira_key, 0) < _JIRA_EFFORT_TTL:
        return _jira_effort_cache[jira_key]

    url = (f'{_JIRA_REST}{jira_key}'
           '?fields=timeoriginalestimate,aggregatetimeoriginalestimate'
           ',subtasks,issuelinks,issuetype')
    try:
        resp = session.get(url, timeout=15)
        if resp.status_code != 200:
            logger.debug('Jira effort %s: HTTP %s', jira_key, resp.status_code)
            _jira_effort_cache[jira_key]    = 0
            _jira_effort_cache_ts[jira_key] = time.time()
            return 0

        fields = resp.json().get('fields', {})

        # Use direct estimate if available
        estimate = (fields.get('timeoriginalestimate') or
                    fields.get('aggregatetimeoriginalestimate') or 0)
        if estimate > 0:
            _jira_effort_cache[jira_key]    = estimate
            _jira_effort_cache_ts[jira_key] = time.time()
            return estimate

        # No direct estimate — sum children recursively
        child_keys = _jira_child_keys(jira_key, fields, session)
        total = sum(
            _jira_effort_seconds(k, session, depth + 1, visited)
            for k in child_keys
        )

        # If this is an Epic, also sum direct estimates of its Agile stories
        # (leaf-only: do NOT recurse into those stories' own children — avoids
        #  pulling in blocker/unrelated estimates)
        if fields.get('issuetype', {}).get('name') == 'Epic':
            try:
                r = session.get(
                    f'{_JIRA_AGILE}{jira_key}/issue'
                    '?fields=timeoriginalestimate&maxResults=100',
                    timeout=15)
                if r.status_code == 200:
                    for iss in r.json().get('issues', []):
                        k = iss.get('key')
                        if k and k not in visited and k not in child_keys:
                            est = (iss.get('fields') or {}).get('timeoriginalestimate') or 0
                            total += est
            except Exception as exc:
                logger.debug('Agile epic children %s: %s', jira_key, exc)
        _jira_effort_cache[jira_key]    = total
        _jira_effort_cache_ts[jira_key] = time.time()
        return total

    except Exception as exc:
        logger.debug('Jira effort fetch %s: %s', jira_key, exc)
        _jira_effort_cache[jira_key]    = 0
        _jira_effort_cache_ts[jira_key] = time.time()
        return 0


def _fetch_row_efforts(rows: list[dict]) -> dict:
    """Return {hsd_id: {xekmd_h, ualkmd_h, umd_sysman_h, total_h}} for every row."""
    all_keys: set = set()
    for row in rows:
        for col in ('xekmd', 'ualkmd', 'umd_sysman'):
            all_keys.update(row.get(col, []))

    now = time.time()
    uncached = [k for k in all_keys
                if k not in _jira_effort_cache or
                now - _jira_effort_cache_ts.get(k, 0) >= _JIRA_EFFORT_TTL]

    if uncached:
        session = _jira_session()
        for key in uncached:
            _jira_effort_seconds(key, session)

    result = {}
    for row in rows:
        hid = row['hsd_id']
        comp: dict = {}
        total_secs = 0
        for col in ('xekmd', 'ualkmd', 'umd_sysman'):
            secs = sum(_jira_effort_cache.get(k, 0) for k in row.get(col, []))
            hours = round(secs / 3600) if secs > 0 else 0
            comp[f'{col}_h'] = hours
            total_secs += secs
        comp['total_h'] = round(total_secs / 3600) if total_secs > 0 else 0
        result[hid] = comp
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# Excel cache coordinator
# ═══════════════════════════════════════════════════════════════════════════════

def _get_groups() -> tuple[list[dict], list[dict], str]:
    """Return (jgs_groups, other_groups, error_msg). Downloads/parses Excel if stale."""
    if _excel_needs_refresh():
        ok, dl_err = _download_excel()
        if not ok and not _EXCEL_PATH.exists():
            _excel_cache.update({'jgs': [], 'other': [], 'error': dl_err})
            return [], [], dl_err
        jgs, other, parse_err = _parse_excel()
        _excel_cache.update({'jgs': jgs, 'other': other, 'ts': time.time(),
                             'error': parse_err or dl_err})
    elif _excel_cache['jgs'] is None:
        jgs, other, parse_err = _parse_excel()
        _excel_cache.update({'jgs': jgs, 'other': other, 'ts': time.time(),
                             'error': parse_err})

    return (_excel_cache.get('jgs') or [],
            _excel_cache.get('other') or [],
            _excel_cache.get('error', ''))


# ═══════════════════════════════════════════════════════════════════════════════
# Routes
# ═══════════════════════════════════════════════════════════════════════════════

@bp.route('/jgs-4plus2')
def jgs_4plus2_index():
    jgs_groups, other_groups, excel_error = _get_groups()
    all_groups = jgs_groups + other_groups
    hsd_ids = [g['hsd_id'] for g in all_groups]
    titles  = _fetch_hsd_titles(hsd_ids)
    edits   = _load_edits()

    def _build_rows(groups):
        rows = []
        for g in groups:
            hid = g['hsd_id']
            e   = edits.get(hid, {})
            ext_raw = e.get('ext_deps', '')
            row = {
                **g,
                'title':       titles.get(hid, ''),
                'eta':         e.get('eta', ''),
                'comments':    e.get('comments', ''),
                'ext_deps':    ext_raw,
                'ext_dep_ids': re.findall(r'[A-Za-z]+-\d+', ext_raw),
                'updated_at':  e.get('updated_at', ''),
            }
            # Merge user-saved Jira IDs with Excel-sourced ones (union, deduped)
            for col in ('xekmd', 'ualkmd', 'umd_sysman'):
                db_val = e.get(f'jira_{col}', '').strip()
                if db_val:
                    db_ids = [j.upper() for j in re.findall(r'[A-Za-z]+-\d+', db_val)]
                    excel_ids = row[col]
                    row[col] = excel_ids + [j for j in db_ids if j not in excel_ids]
            rows.append(row)
        return rows

    jgs_rows   = _build_rows(jgs_groups)
    other_rows = _build_rows(other_groups)

    # Fetch Jira effort per component (uses in-memory cache, 4 h TTL)
    all_rows = jgs_rows + other_rows
    efforts  = _fetch_row_efforts(all_rows)
    for row in all_rows:
        e = efforts.get(row['hsd_id'], {})
        row['effort_xekmd']      = e.get('xekmd_h', 0)
        row['effort_ualkmd']     = e.get('ualkmd_h', 0)
        row['effort_umd_sysman'] = e.get('umd_sysman_h', 0)
        row['effort_total']      = e.get('total_h', 0)

    last_refresh = ''
    if _EXCEL_TS.exists():
        try:
            ts = float(_EXCEL_TS.read_text().strip())
            last_refresh = datetime.fromtimestamp(ts).strftime('%d %b %Y %H:%M')
        except Exception:
            pass

    return render_template(
        'jgs_4plus2.html',
        jgs_rows=jgs_rows,
        other_rows=other_rows,
        excel_error=excel_error,
        last_refresh=last_refresh,
        hsd_app_url=_HSD_APP,
        jira_base=_JIRA_BASE,
    )


@bp.route('/jgs-4plus2/save', methods=['POST'])
def jgs_4plus2_save():
    body  = request.get_json(silent=True) or {}
    hid   = str(body.get('hsd_id', '')).strip()[:20]
    field = str(body.get('field', '')).strip()
    value = str(body.get('value', ''))[:2000]
    if not hid or field not in _VALID_FIELDS:
        return jsonify({'error': f'bad request — field must be one of {sorted(_VALID_FIELDS)}'}), 400
    try:
        _save_edit(hid, field, value)
        return jsonify({'ok': True})
    except Exception as exc:
        logger.error('jgs_4plus2 save error: %s', exc)
        return jsonify({'error': str(exc)}), 500


@bp.route('/jgs-4plus2/refresh-excel', methods=['POST'])
def jgs_4plus2_refresh_excel():
    # Invalidate caches so next GET triggers a fresh download
    _excel_cache.update({'jgs': None, 'other': None, 'ts': 0.0, 'error': ''})
    _jira_effort_cache.clear()
    _jira_effort_cache_ts.clear()
    try:
        _EXCEL_TS.unlink(missing_ok=True)
    except Exception:
        pass
    ok, err = _download_excel()
    if ok:
        _excel_cache.update({'jgs': None, 'other': None})  # force re-parse on next request
    return jsonify({'ok': ok, 'error': err})


@bp.route('/jgs-4plus2/api/data')
def jgs_4plus2_api_data():
    """Return edits merged with Excel Jiras (for 60-second polling).

    Jira fields are merged so Excel-sourced IDs are never dropped by polling.
    """
    edits = _load_edits()
    try:
        jgs_groups, other_groups, _ = _get_groups()
        excel_jiras: dict = {}
        for g in jgs_groups + other_groups:
            hid = g['hsd_id']
            excel_jiras[hid] = {col: g.get(col, []) for col in ('xekmd', 'ualkmd', 'umd_sysman')}
    except Exception:
        excel_jiras = {}

    result: dict = {}
    for hid, e in edits.items():
        ex  = excel_jiras.get(hid, {})
        row = dict(e)
        for col in ('xekmd', 'ualkmd', 'umd_sysman'):
            db_val   = e.get(f'jira_{col}', '').strip()
            xl_ids   = ex.get(col, [])
            db_ids   = [j.upper() for j in re.findall(r'[A-Za-z]+-\d+', db_val)] if db_val else []
            merged   = xl_ids + [j for j in db_ids if j not in xl_ids]
            row[f'jira_{col}'] = ' '.join(merged)
        result[hid] = row

    # Also include HSDs that have Excel Jiras but no DB entry
    for hid, ex in excel_jiras.items():
        if hid not in result:
            merged_row: dict = {}
            for col in ('xekmd', 'ualkmd', 'umd_sysman'):
                merged_row[f'jira_{col}'] = ' '.join(ex.get(col, []))
            result[hid] = merged_row

    return jsonify(result)
